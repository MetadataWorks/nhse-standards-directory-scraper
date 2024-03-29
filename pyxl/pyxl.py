
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook  
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, PatternFill
#importing pandas as pd
import pandas as pd




thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

headers = {'A':"Persistent Identifer", 'B':"Summary", 'P':"Documentation", 'U':"Publisher", 'Z':"Document Control", 'AN':"Data Status"}



def create_excel(api_data, keys):

    work_book = Workbook()  
    print('Workbook')

    work_sheet = work_book.active  
    print('worksheet')

    work_sheet.append(keys)
    print('inserted headers')

    for col in range(1, 42):
        work_sheet.cell(1, col).alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)

    for x in api_data:
        work_sheet.append(list(x.values()))
    
    print('inserted data')


    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            if cell.value:
                colLetter = get_column_letter(cell.column)
                dims[colLetter] = max((dims.get(colLetter, 0), len(str(cell.value))))    

    for col, value in dims.items():
        if value<10:
            value=10
        work_sheet.column_dimensions[col].width = value

    print('width settled')
    

    work_sheet.insert_rows(idx=1)
    work_sheet.insert_rows(idx=1)

    print('new rows inserted')


    # cells to merge  
    work_sheet.merge_cells('A1:A2') 
    work_sheet.merge_cells('B1:O2') 
    work_sheet.merge_cells('P1:T2') 
    work_sheet.merge_cells('U1:Y2') 
    work_sheet.merge_cells('Z1:AM1')
    work_sheet.merge_cells('Z2:AL2')
    work_sheet.merge_cells('AN1:AO2') 

    print('cells merged')


    for col in ['A', 'B', 'P', 'U', 'Z', 'AN']:
        work_sheet[f'{col}1'].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
        work_sheet[f'{col}1'].border = thin_border
        work_sheet[f'{col}1'].fill = PatternFill(start_color='f1c232', fill_type='solid')
        work_sheet[f'{col}1'].value = headers[col]
    

    work_sheet['Z2'].fill = PatternFill(start_color='f1c232', fill_type='solid')
    work_sheet['Z2'].border = thin_border

    work_sheet['AM2'].fill = PatternFill(start_color='e69138', fill_type='solid')
    work_sheet['AM2'].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    work_sheet['AM2'].border = thin_border
    work_sheet['AM2'].value = 'Reviews'


    print('rows formatted')


    # save the workbook  
    work_book.save('scraped_data/NHS_API_CATALOGUE.xlsx')
    print('xlsx saved')


    

    read_file = pd.read_excel("scraped_data/NHS_API_CATALOGUE.xlsx")
    print('xlsx read')
    

    read_file.to_csv ("scraped_data/NHS_API_CATALOGUE.csv", index=None)
    print('csv saved')
